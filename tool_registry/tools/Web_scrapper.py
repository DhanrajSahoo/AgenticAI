from crewai.tools import BaseTool
from typing import Type
from pydantic import BaseModel, Field
import os
import ssl
import json
import re
import requests
from bs4 import BeautifulSoup
from googlesearch import search
from openpyxl import load_workbook

# Skip SSL validation (optional, useful if you hit HTTPS issues)
ssl._create_default_https_context = ssl._create_unverified_context


# --------- Gather Info Function ---------
def gather_info(query) -> list:
    url_data = []
    urls = []

    try:
        for url in search(query, num_results=10):  # Get more to filter .pdf
            if url.lower().endswith(".pdf"):
                continue  # ✅ Skip PDF links
            if url.startswith("http"):
                urls.append(url)
            if len(urls) >= 5:  # Keep only first 5 non-PDF links
                break

        for url in urls:
            try:
                response = requests.get(url, timeout=5)
                soup = BeautifulSoup(response.text, 'html.parser')
                data = {'url': url, 'data': []}

                for paragraph in soup.find_all('p'):
                    para = paragraph.get_text(strip=True)
                    if para:
                        cleaned = re.sub(r'\n\s+', r' ', para)
                        cleaned = re.sub(r'[\u00a0-\u00af]', r' ', cleaned)
                        data['data'].append(cleaned)
                url_data.append(data)

            except Exception as e:
                url_data.append({'url': url, 'error': str(e)})
        return url_data

    except Exception as e:
        return [{"error": str(e)}]


# --------- Web Scraper Function ---------
def web_scrapper(file_path):
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        keys = [str(item).strip().upper() for item in rows[0]]
        result_data = []

        if "NAME" in keys or "CONTACT NAME" in keys:
            for row in rows[1:]:
                if not any(row):
                    continue

                response_data = {}
                items = [str(item).strip() if item else "" for item in row]
                query = " ".join(items)
                existing_data = dict(zip(keys, items))

                fetched_data = gather_info(query)

                response_data["Existing data"] = existing_data
                response_data["New data"] = fetched_data
                result_data.append(response_data)

            return result_data

        else:
            return {"error": "Excel must have a 'Name' or 'Contact Name' column."}

    except Exception as e:
        return {"error": f"Processing failed: {str(e)}"}


# --------- Pydantic Schema ---------
class WebscrappingSchema(BaseModel):
    file_path: str = Field(..., description="Path to Excel file with contact data.")

    class Config:
        allow_population_by_field_name = True
        allow_population_by_alias = True


# --------- Tool Class ---------
class WebScrappingTool(BaseTool):
    name: str = "Web Scraper and Comparison Tool"
    description: str = "Reads an Excel, fetches latest web data, and compares."
    args_schema: Type[BaseModel] = WebscrappingSchema

    def _run(self, file_path: str) -> str:
        try:
            result = web_scrapper(file_path)
            return json.dumps(result, indent=2)
        except Exception as e:
            return json.dumps({"error": f"Something went wrong: {str(e)}"})

    def run(self, input_data: WebscrappingSchema) -> str:
        return self._run(input_data.file_path)
